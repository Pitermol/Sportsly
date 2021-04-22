from numpy import exp, array, random, dot, reshape
from openpyxl import load_workbook


class NeuralNetwork:
    def __init__(self, weights=[]):
        # Seed the random number generator, so it generates the same numbers
        # every time the program runs.
        random.seed(1)

        # We model a single neuron, with 3 input connections and 1 output connection.
        # We assign random weights to a 3 x 1 matrix, with values in the range -1 to 1
        # and mean 0.
        if len(weights) == 0:
            self.synaptic_weights = 2 * random.random((3, 1)) - 1
        else:
            self.synaptic_weights = weights
        #print(self.synaptic_weights)

    # The Sigmoid function, which describes an S shaped curve.
    # We pass the weighted sum of the inputs through this function to
    # normalise them between 0 and 1.
    def __sigmoid(self, x):
        return 1 / (1 + exp(-x))

    # The derivative of the Sigmoid function.
    # This is the gradient of the Sigmoid curve.
    # It indicates how confident we are about the existing weight.
    def __sigmoid_derivative(self, x):
        return x * (1 - x)

    # We train the neural network through a process of trial and error.
    # Adjusting the synaptic weights each time.
    def train(self, training_set_inputs, training_set_outputs, number_of_training_iterations):
        for iteration in range(number_of_training_iterations):
            # Pass the training set through our neural network (a single neuron).
            output = self.think(training_set_inputs)

            # Calculate the error (The difference between the desired output
            # and the predicted output).

            error = training_set_outputs - output

            # Multiply the error by the input and again by the gradient of the Sigmoid curve.
            # This means less confident weights are adjusted more.
            # This means inputs, which are zero, do not cause changes to the weights.
            #print(training_set_inputs.T)
            adjustment = dot(reshape(training_set_inputs.T, (3,1)), reshape(error * self.__sigmoid_derivative(output), (1,1)))

            # Adjust the weights.
            self.synaptic_weights += adjustment

    # The neural network thinks.
    def think(self, inputs):
        # Pass inputs through our neural network (our single neuron).
        return self.__sigmoid(dot(inputs, self.synaptic_weights))


if __name__ == "__main__":

    #Intialise a single neuron neural network.
    neural_network = NeuralNetwork()

#    print("Random starting synaptic weights: ")
#    print (neural_network.synaptic_weights)

    # The training set. We have 4 examples, each consisting of 3 input values
    # and 1 output value.

    wb = load_workbook('./hockey_data.xlsx')
    sheet = wb['Sheet']
    inputs_arr = []
    outputs = []
    start = 3
    end = 153
    for i in range(start, end, 2):
        one = []
        for j in range(2, 5):
            #print(i, j)
            razn = float(sheet.cell(row=i, column=j).value) - float(sheet.cell(row=i+1, column=j).value)
            one.append(razn)
        #print(sheet.cell(row=i, column=5).value)
        result = str(sheet.cell(row=i, column=5).value).split(":")
        if result[0] == result[1]:
            if result[2] == "00":
                outputs.append(0.3)
            else:
                outputs.append(0.7)
        else:
            outputs.append(float(1.0 * int(int(result[0]) < int(result[1]))))
        #outputs.append(float(sheet.cell(row=i+1, column=5).value))
        inputs_arr.append(one)
    training_set_outputs = array(outputs).T
    training_set_inputs = array(inputs_arr)
    # Train the neural network using a training set.
    # Do it 10,000 times and make small adjustments each time.
    for i in range(len(training_set_inputs)):
        #print(training_set_inputs[i])
        #print(training_set_outputs[i])
        neural_network.train(training_set_inputs[i], training_set_outputs[i], 15)

    with open("weights", "w") as file:
        for weight in neural_network.synaptic_weights:
            file.write("".join(str(weight[0])) + "\n")
    file.close()

    print("New synaptic weights after training: ")
    print(neural_network.synaptic_weights)

    # Test the neural network with a new situation.
    print("new situation")
    print(neural_network.think(array([-0.2, -0.51, -0.16])))